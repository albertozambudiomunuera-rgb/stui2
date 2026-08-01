#!/usr/bin/env python3
"""
agrupar-respuestas.py
──────────────────────────────────────────────────────────────────────────────
Agrupa los cuestionarios Excel devueltos por los urólogos y genera un
libro de análisis con las puntuaciones ya calculadas.

USO
    1. Guarda todos los archivos devueltos en una carpeta, p. ej.  ./respuestas/
    2. python3 agrupar-respuestas.py ./respuestas/
    3. Genera  STUIapp_Analisis_Usabilidad.xlsx  en la misma carpeta

Calcula por evaluador: SUS (0-100), medias uMARS por dominio, utilidad
clínica y finalización de tareas. Y en agregado: media, DE, mediana, rango,
desglose por centro y comparación con la norma SUS de 68.

No depende de que las fórmulas del archivo estén intactas: lee las respuestas
en bruto y recalcula todo en Python. Así da igual si alguien rompió una celda.
──────────────────────────────────────────────────────────────────────────────
"""

import sys, os, glob, statistics as st
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

F = "Arial"
H1 = Font(name=F, size=14, bold=True, color="FFFFFF")
B  = Font(name=F, size=10, bold=True)
N  = Font(name=F, size=10)
NI = Font(name=F, size=9, italic=True, color="595959")
HDR = PatternFill("solid", fgColor="1F3864")
SUB = PatternFill("solid", fgColor="D9E2F3")
CALC = PatternFill("solid", fgColor="E2EFDA")
thin = Side(style="thin", color="BFBFBF")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
CTR = Alignment(horizontal="center", vertical="center")

UMARS_BLOCKS = [
    ("Engagement", 0, 4), ("Funcionalidad", 5, 8), ("Estetica", 9, 11),
    ("Informacion", 12, 15), ("Subjetiva", 16, 19), ("Impacto", 20, 25),
]


def num(v):
    """Convierte a número si se puede; si no, None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip().replace(",", "."))
    except ValueError:
        return None


def sus_score(vals):
    """SUS estándar: impares -1, pares 5-x, suma x 2.5. Requiere los 10 ítems."""
    if len(vals) != 10 or any(v is None for v in vals):
        return None
    total = 0.0
    for i, v in enumerate(vals):
        total += (v - 1) if i % 2 == 0 else (5 - v)
    return round(total * 2.5, 1)


def leer(path):
    """Extrae un evaluador desde su archivo. Devuelve dict o None si ilegible."""
    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        print(f"  !! No se pudo abrir {os.path.basename(path)}: {e}")
        return None

    d = {"archivo": os.path.basename(path)}

    def sheet(name):
        for s in wb.sheetnames:
            if s.strip().lower().startswith(name.lower()):
                return wb[s]
        return None

    # ── Datos del evaluador ──
    ws = sheet("1.")
    campos = ["centro", "anios", "uso_previo", "dias_uso", "dispositivo",
              "navegador", "fecha"]
    for i, k in enumerate(campos):
        d[k] = ws.cell(4 + i, 3).value if ws else None

    # ── SUS ──
    ws = sheet("2.")
    sus = [num(ws.cell(7 + i, 4).value) for i in range(10)] if ws else [None] * 10
    d["sus_items"] = sus
    d["SUS"] = sus_score(sus)

    # ── uMARS ──
    ws = sheet("3.")
    um = [num(ws.cell(7 + i, 4).value) for i in range(26)] if ws else [None] * 26
    d["umars_items"] = um
    for name, a, b in UMARS_BLOCKS:
        vals = [v for v in um[a:b + 1] if v is not None]
        d[f"uMARS_{name}"] = round(st.mean(vals), 2) if vals else None
    obj = [v for v in um[0:16] if v is not None]
    d["uMARS_global"] = round(st.mean(obj), 2) if obj else None

    # ── Utilidad clínica ──
    ws = sheet("4.")
    ut = [num(ws.cell(6 + i, 4).value) for i in range(8)] if ws else [None] * 8
    d["util_items"] = ut
    vals = [v for v in ut if v is not None]
    d["Utilidad"] = round(st.mean(vals), 2) if vals else None

    # Abiertas: bloques fusionados a partir de la fila 19, cada 4 filas
    ab = []
    if ws:
        for i in range(5):
            ab.append(ws.cell(19 + i * 4, 2).value)
    d["abiertas"] = ab

    # ── Tareas ──
    ws = sheet("5.")
    if ws:
        d["tareas"] = [ws.cell(4 + i, 3).value for i in range(5)]
        tiempos = [num(ws.cell(4 + i, 4).value) for i in range(5)]
        tiempos = [t for t in tiempos if t is not None]
        d["tiempo_total"] = sum(tiempos) if tiempos else None
        d["tareas_ok"] = sum(1 for t in d["tareas"]
                             if t and "sin ayuda" in str(t).lower())
    else:
        d["tareas"], d["tiempo_total"], d["tareas_ok"] = [None] * 5, None, None

    return d


def desc(vals):
    """Estadística descriptiva. Devuelve dict con n, media, DE, mediana, min, max."""
    v = [x for x in vals if x is not None]
    if not v:
        return None
    return {
        "n": len(v),
        "media": round(st.mean(v), 2),
        "de": round(st.stdev(v), 2) if len(v) > 1 else 0.0,
        "mediana": round(st.median(v), 2),
        "min": round(min(v), 2),
        "max": round(max(v), 2),
    }


def main():
    carpeta = sys.argv[1] if len(sys.argv) > 1 else "."
    files = sorted(glob.glob(os.path.join(carpeta, "*.xlsx")))
    files = [f for f in files
             if "Analisis" not in f and not os.path.basename(f).startswith("~$")]

    if not files:
        print(f"No se encontraron archivos .xlsx en {carpeta}")
        sys.exit(1)

    print(f"Encontrados {len(files)} archivos\n")
    evals = []
    for f in files:
        d = leer(f)
        if d:
            evals.append(d)
            print(f"  OK  {d['archivo'][:44]:46} "
                  f"SUS={d['SUS'] if d['SUS'] is not None else '—'}")

    if not evals:
        print("Ningún archivo legible.")
        sys.exit(1)

    # ══════════════════ Libro de análisis ══════════════════
    wb = Workbook()

    # ── Hoja 1: datos individuales ──
    ws = wb.active
    ws.title = "Individual"
    cols = ["Archivo", "Centro", "Años", "Uso previo", "Dispositivo",
            "SUS", "uMARS global", "Engagement", "Funcionalidad", "Estética",
            "Información", "Subjetiva", "Impacto", "Utilidad",
            "Tareas s/ayuda", "Tiempo (min)"]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    c = ws.cell(1, 1, f"STUIapp · Resultados individuales · n={len(evals)}")
    c.font = H1; c.fill = HDR; c.alignment = CTR
    ws.row_dimensions[1].height = 26

    for j, h in enumerate(cols):
        c = ws.cell(3, j + 1, h); c.font = B; c.fill = SUB; c.border = BOX
        c.alignment = Alignment(wrap_text=True, horizontal="center")
    ws.column_dimensions["A"].width = 34
    for j in range(2, len(cols) + 1):
        ws.column_dimensions[chr(64 + j) if j <= 26 else "A"].width = 13

    for i, d in enumerate(evals):
        r = 4 + i
        vals = [d["archivo"], d.get("centro"), d.get("anios"),
                d.get("uso_previo"), d.get("dispositivo"),
                d["SUS"], d["uMARS_global"],
                d["uMARS_Engagement"], d["uMARS_Funcionalidad"],
                d["uMARS_Estetica"], d["uMARS_Informacion"],
                d["uMARS_Subjetiva"], d["uMARS_Impacto"],
                d["Utilidad"], d.get("tareas_ok"), d.get("tiempo_total")]
        for j, v in enumerate(vals):
            c = ws.cell(r, j + 1, v); c.font = N; c.border = BOX
            if j >= 5:
                c.alignment = CTR

    # ── Hoja 2: descriptivos ──
    ws = wb.create_sheet("Descriptivos")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    c = ws.cell(1, 1, "Estadística descriptiva")
    c.font = H1; c.fill = HDR; c.alignment = CTR
    ws.column_dimensions["A"].width = 26
    for col in "BCDEFG":
        ws.column_dimensions[col].width = 12

    heads = ["Variable", "n", "Media", "DE", "Mediana", "Mín", "Máx"]
    for j, h in enumerate(heads):
        c = ws.cell(3, j + 1, h); c.font = B; c.fill = SUB; c.border = BOX
        c.alignment = CTR

    metrics = [("SUS (0-100)", "SUS"), ("uMARS global", "uMARS_global")]
    metrics += [(f"uMARS {n}", f"uMARS_{n}") for n, _, _ in UMARS_BLOCKS]
    metrics += [("Utilidad clínica", "Utilidad"),
                ("Tareas sin ayuda (0-5)", "tareas_ok"),
                ("Tiempo caso (min)", "tiempo_total")]

    r = 4
    for label, key in metrics:
        s = desc([d.get(key) for d in evals])
        ws.cell(r, 1, label).font = B; ws.cell(r, 1).border = BOX
        if s:
            for j, k in enumerate(["n", "media", "de", "mediana", "min", "max"]):
                c = ws.cell(r, j + 2, s[k]); c.font = N; c.border = BOX
                c.alignment = CTR
        r += 1

    # Interpretación SUS
    s = desc([d["SUS"] for d in evals])
    if s:
        r += 1
        ws.cell(r, 1, "INTERPRETACIÓN SUS").font = B
        ws.cell(r, 1).fill = CALC; r += 1
        m = s["media"]
        adj = ("Excelente (percentil A)" if m > 80.3 else
               "Bueno (percentil B-C)" if m >= 68 else
               "Aceptable justo (D)" if m >= 51 else "Pobre (F)")
        for k, v in [("Media obtenida", m),
                     ("Norma histórica", 68),
                     ("Diferencia", round(m - 68, 1)),
                     ("Categoría", adj)]:
            ws.cell(r, 1, k).font = N; ws.cell(r, 1).border = BOX
            c = ws.cell(r, 2, v); c.font = B; c.border = BOX; c.alignment = CTR
            r += 1
        r += 1
        ws.cell(r, 1, "El SUS no es un porcentaje: 68 es el percentil 50, "
                      "no un 68%.").font = NI

    # ── Hoja 3: por centro ──
    ws = wb.create_sheet("Por centro")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    c = ws.cell(1, 1, "Resultados por centro")
    c.font = H1; c.fill = HDR; c.alignment = CTR
    ws.column_dimensions["A"].width = 24
    for col in "BCDEF":
        ws.column_dimensions[col].width = 14

    centros = sorted({str(d.get("centro")) for d in evals if d.get("centro")})
    heads = ["Centro", "n", "SUS medio", "DE", "uMARS", "Utilidad"]
    for j, h in enumerate(heads):
        c = ws.cell(3, j + 1, h); c.font = B; c.fill = SUB; c.border = BOX
        c.alignment = CTR
    for i, cen in enumerate(centros):
        sub = [d for d in evals if str(d.get("centro")) == cen]
        ss = desc([d["SUS"] for d in sub])
        su = desc([d["uMARS_global"] for d in sub])
        sut = desc([d["Utilidad"] for d in sub])
        vals = [cen, len(sub),
                ss["media"] if ss else None, ss["de"] if ss else None,
                su["media"] if su else None, sut["media"] if sut else None]
        for j, v in enumerate(vals):
            c = ws.cell(4 + i, j + 1, v); c.font = N; c.border = BOX
            if j: c.alignment = CTR
    r = 5 + len(centros)
    ws.cell(r, 1, "Con n≈5 por centro, describe pero NO hagas contrastes "
                  "inferenciales entre centros.").font = NI

    # ── Hoja 4: respuestas abiertas ──
    ws = wb.create_sheet("Abiertas")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    c = ws.cell(1, 1, "Respuestas abiertas · material para análisis temático")
    c.font = H1; c.fill = HDR; c.alignment = CTR
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 90

    preguntas = ["Qué funciona peor", "Qué añadirías",
                 "Errores detectados", "Barrera principal", "Otros comentarios"]
    for j, h in enumerate(["Evaluador", "Pregunta", "Respuesta"]):
        c = ws.cell(3, j + 1, h); c.font = B; c.fill = SUB; c.border = BOX
    r = 4
    for d in evals:
        for i, q in enumerate(preguntas):
            txt = d["abiertas"][i] if i < len(d["abiertas"]) else None
            if txt and str(txt).strip():
                ws.cell(r, 1, d["archivo"][:20]).font = N
                ws.cell(r, 2, q).font = N
                c = ws.cell(r, 3, str(txt)); c.font = N
                c.alignment = Alignment(wrap_text=True, vertical="top")
                for j in range(1, 4):
                    ws.cell(r, j).border = BOX
                r += 1

    out = os.path.join(carpeta, "STUIapp_Analisis_Usabilidad.xlsx")
    wb.save(out)

    print(f"\n{'='*54}")
    if s:
        print(f"  SUS medio:  {s['media']}  (DE {s['de']}, n={s['n']})")
        print(f"  Norma:      68")
    print(f"  Guardado:   {out}")
    print("="*54)


if __name__ == "__main__":
    main()
